import io
import os
import tempfile
import urllib.parse

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel

import pandas as pd
import psycopg
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from importer import read_and_normalize_excel, map_row, upsert_project, upsert_snapshot, compute_deltas

app = FastAPI()

# Static + templates
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# DB
DB_DSN = os.environ.get("DB_DSN", "postgresql://postgres:TU_PASSWORD@localhost:5432/mecalux")


@app.on_event("startup")
def startup_init() -> None:
    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            ensure_historical_storage(cur)
            ensure_project_tasks_storage(cur)
        conn.commit()

PHASES = ("design", "development", "pem", "hypercare")
ROLES = ("pm", "consultant", "technician")


class AssignedHoursPhaseIn(BaseModel):
    phase: str
    hours: float | None = None


class AssignedHoursRoleIn(BaseModel):
    role: str
    hours: float | None = None


class ProjectCommentIn(BaseModel):
    comment_text: str | None = None


TASK_TYPES = {"TASK", "PP"}
TASK_OWNER_ROLES = {"PM", "CONSULTORIA", "TECH", "COMERCIAL", "CLIENTE"}
TASK_STATUSES = {"OPEN", "IN_PROGRESS", "PAUSED", "CLOSED"}


class ProjectTaskCreateIn(BaseModel):
    project_id: int
    type: str
    owner_role: str
    planned_date: str | None = None
    status: str = "OPEN"
    description: str


class ProjectTaskStatusIn(BaseModel):
    status: str


class ProjectTaskUpdateIn(BaseModel):
    type: str
    owner_role: str
    planned_date: str | None = None
    status: str
    description: str


def normalize_comment(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, memoryview):
        value = value.tobytes()
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="ignore")
    return str(value)


def to_float(value: object) -> float | None:
    if value is None:
        return None
    return float(value)


def to_date_iso(value: object) -> str | None:
    if value is None:
        return None
    return value.isoformat()


def fetch_deviations_results(
    teams: list[str],
    phases: list[str],
) -> tuple[list[dict], list[str], set[str], list[str | None]]:
    where_clauses = []
    params: list[object] = []
    if teams:
        where_clauses.append("s.team = ANY(%s::text[])")
        params.append(teams)
    if phases:
        where_clauses.append("s.order_phase = ANY(%s::text[])")
        params.append(phases)
    where_sql = "WHERE COALESCE(p.is_historical, FALSE) = FALSE"
    if where_clauses:
        where_sql += " AND " + " AND ".join(where_clauses)

    sql = f"""
    WITH ranked AS (
        SELECT
            p.id AS project_id,
            p.project_code,
            p.project_name,
            s.team,
            s.order_phase,
            s.ordered_total,
            s.real_hours,
            s.desviacion_pct,
            s.progress_w,
            s.report_date,
            s.comments,
            s.snapshot_year,
            s.snapshot_week,
            s.snapshot_at,
            ROW_NUMBER() OVER (
                PARTITION BY p.id
                ORDER BY s.snapshot_year DESC, s.snapshot_week DESC, s.snapshot_at DESC
            ) AS rn
        FROM projects p
        JOIN project_snapshot s ON s.project_id = p.id
        {where_sql}
    ),
    eligible_projects AS (
        SELECT project_id
        FROM ranked
        WHERE rn <= 2
        GROUP BY project_id
        HAVING BOOL_OR(COALESCE(desviacion_pct, 0) > 0)
    )
    SELECT r.*
    FROM ranked r
    JOIN eligible_projects e ON e.project_id = r.project_id
    WHERE r.rn <= 5
    """
    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()
            colnames = [desc[0] for desc in cur.description]

    grouped: dict[int, list[dict]] = {}
    for row in rows:
        record = dict(zip(colnames, row))
        grouped.setdefault(record["project_id"], []).append(record)

    columns = [
        "Proyecto",
        "Código",
        "Equipo",
        "Order phase",
        "Fecha",
        "H.Total",
        "H.Real",
    ]

    latest_snapshot = None
    for items in grouped.values():
        for item in items:
            key = (item.get("snapshot_year") or 0, item.get("snapshot_week") or 0)
            if latest_snapshot is None or key > latest_snapshot:
                latest_snapshot = key

    def prev_week_label(week: int, offset: int) -> str:
        week_value = week - offset
        while week_value <= 0:
            week_value += 52
        return f"W{week_value:02d}"

    start_week = latest_snapshot[1] if latest_snapshot and latest_snapshot[1] else 5
    snapshot_labels = [prev_week_label(int(start_week), idx) for idx in range(5)]
    snapshot_deviation_columns = [f"{label} Desviación" for label in snapshot_labels]
    snapshot_real_columns = [f"{label} H.Real" for label in snapshot_labels]
    snapshot_progress_columns = [f"{label} Avance" for label in snapshot_labels]
    for deviation_col, real_col, progress_col in zip(
        snapshot_deviation_columns,
        snapshot_real_columns,
        snapshot_progress_columns,
    ):
        columns.append(deviation_col)
        columns.append(real_col)
        columns.append(progress_col)
    columns.append("Comentario")
    numeric_columns = set(columns) - {"Proyecto", "Código", "Equipo", "Order phase", "Fecha", "Comentario"}

    results = []
    row_styles: list[str | None] = []
    for items in grouped.values():
        items_sorted = sorted(items, key=lambda item: item["rn"])
        latest = next((item for item in items_sorted if item["rn"] == 1), None)
        prev = next((item for item in items_sorted if item["rn"] == 2), None)
        if latest is None or prev is None:
            continue
        latest_dev = to_float(latest.get("desviacion_pct"))
        prev_dev = to_float(prev.get("desviacion_pct"))
        row_style = None
        if latest_dev is not None and prev_dev is not None:
            if latest_dev > prev_dev:
                row_style = "danger"
            elif latest_dev == prev_dev:
                row_style = "warning"
            else:
                row_style = "success"
        row = {
            "Proyecto": latest.get("project_name"),
            "Código": latest.get("project_code"),
            "Equipo": latest.get("team"),
            "Order phase": latest.get("order_phase"),
            "Fecha": latest.get("report_date"),
            "H.Total": to_float(latest.get("ordered_total")),
            "H.Real": to_float(latest.get("real_hours")),
        }

        by_label: dict[str, dict[str, float | None]] = {}
        for snapshot in items_sorted:
            snapshot_week = snapshot.get("snapshot_week")
            if snapshot_week is None:
                continue
            try:
                label = f"W{int(snapshot_week):02d}"
            except (TypeError, ValueError):
                continue
            if label not in by_label:
                by_label[label] = {
                    "deviation": to_float(snapshot.get("desviacion_pct")),
                    "real": to_float(snapshot.get("real_hours")),
                    "progress": to_float(snapshot.get("progress_w")),
                }

        for label, deviation_col, real_col, progress_col in zip(
            snapshot_labels,
            snapshot_deviation_columns,
            snapshot_real_columns,
            snapshot_progress_columns,
        ):
            weekly_values = by_label.get(label) or {}
            row[deviation_col] = weekly_values.get("deviation")
            row[real_col] = weekly_values.get("real")
            row[progress_col] = weekly_values.get("progress")
        row["Comentario"] = normalize_comment(latest.get("comments"))
        results.append(row)
        row_styles.append(row_style)

    sorted_pairs = sorted(
        zip(results, row_styles),
        key=lambda item: (item[0].get("Proyecto") or "").lower(),
    )
    if sorted_pairs:
        results = [pair[0] for pair in sorted_pairs]
        row_styles = [pair[1] for pair in sorted_pairs]
    else:
        results = []
        row_styles = []
    return results, columns, numeric_columns, row_styles


def fetch_filter_options() -> tuple[list[str], list[str]]:
    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT DISTINCT team
                FROM project_snapshot
                WHERE team IS NOT NULL AND team <> ''
                ORDER BY team
                """
            )
            teams = [row[0] for row in cur.fetchall()]
            cur.execute(
                """
                SELECT DISTINCT order_phase
                FROM project_snapshot
                WHERE order_phase IS NOT NULL AND order_phase <> ''
                ORDER BY order_phase
                """
            )
            phases = [row[0] for row in cur.fetchall()]
    return teams, phases


def ensure_details_columns(cur: psycopg.Cursor) -> None:
    cur.execute(
        """
        ALTER TABLE projects
        ADD COLUMN IF NOT EXISTS comments TEXT,
        ADD COLUMN IF NOT EXISTS hours_design NUMERIC DEFAULT 0,
        ADD COLUMN IF NOT EXISTS hours_development NUMERIC DEFAULT 0,
        ADD COLUMN IF NOT EXISTS hours_pem NUMERIC DEFAULT 0,
        ADD COLUMN IF NOT EXISTS hours_hypercare NUMERIC DEFAULT 0,
        ADD COLUMN IF NOT EXISTS hours_pm NUMERIC DEFAULT 0,
        ADD COLUMN IF NOT EXISTS hours_consultant NUMERIC DEFAULT 0,
        ADD COLUMN IF NOT EXISTS hours_technician NUMERIC DEFAULT 0;
        """
    )


def ensure_historical_storage(cur: psycopg.Cursor) -> None:
    cur.execute(
        """
        ALTER TABLE projects
        ADD COLUMN IF NOT EXISTS is_historical BOOLEAN NOT NULL DEFAULT FALSE;
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS projects_historical (
            id BIGSERIAL PRIMARY KEY,
            project_code TEXT NOT NULL UNIQUE,
            project_name TEXT,
            client TEXT,
            company TEXT,
            team TEXT,
            project_manager TEXT,
            consultant TEXT,
            status TEXT,
            moved_to_historical_week TEXT NOT NULL,
            progress_w NUMERIC NOT NULL DEFAULT 100,
            ordered_total NUMERIC,
            real_hours NUMERIC,
            desviacion_pct NUMERIC,
            moved_to_historical_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            last_import_filename TEXT
        );
        """
    )


    cur.execute(
        """
        ALTER TABLE projects_historical
        ADD COLUMN IF NOT EXISTS progress_w NUMERIC NOT NULL DEFAULT 100;
        """
    )

    cur.execute(
        """
        ALTER TABLE projects_historical
        ADD COLUMN IF NOT EXISTS ordered_total NUMERIC,
        ADD COLUMN IF NOT EXISTS real_hours NUMERIC,
        ADD COLUMN IF NOT EXISTS desviacion_pct NUMERIC;
        """
    )

    cur.execute(
        """
        UPDATE projects_historical h
        SET ordered_total = s.ordered_total,
            real_hours = s.real_hours,
            desviacion_pct = s.desviacion_pct
        FROM projects p
        JOIN LATERAL (
            SELECT ordered_total, real_hours, desviacion_pct
            FROM project_snapshot
            WHERE project_id = p.id
            ORDER BY snapshot_year DESC, snapshot_week DESC, snapshot_at DESC
            LIMIT 1
        ) s ON TRUE
        WHERE p.project_code = h.project_code
          AND (h.ordered_total IS NULL OR h.real_hours IS NULL OR h.desviacion_pct IS NULL);
        """
    )


def ensure_project_tasks_storage(cur: psycopg.Cursor) -> None:
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS project_tasks (
            id BIGSERIAL PRIMARY KEY,
            project_id BIGINT NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
            type TEXT NOT NULL CHECK (type IN ('TASK', 'PP')),
            owner_role TEXT NOT NULL CHECK (owner_role IN ('PM', 'CONSULTORIA', 'TECH', 'COMERCIAL', 'CLIENTE')),
            planned_date DATE,
            status TEXT NOT NULL CHECK (status IN ('OPEN', 'IN_PROGRESS', 'PAUSED', 'CLOSED')) DEFAULT 'OPEN',
            description TEXT NOT NULL,
            created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            updated_at TIMESTAMPTZ NOT NULL DEFAULT now()
        );
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_project_tasks_project_status
        ON project_tasks (project_id, status);
        """
    )


def historical_week_label(snapshot_year: int, snapshot_week: int) -> str:
    return f"{snapshot_year}-W{snapshot_week:02d}"


def move_project_to_historical(
    cur: psycopg.Cursor,
    project_id: int,
    project_fields: dict,
    moved_to_historical_week: str,
    filename: str,
) -> None:
    cur.execute(
        """
        UPDATE projects
        SET is_historical = TRUE
        WHERE id = %s
        """,
        (project_id,),
    )
    cur.execute(
        """
        SELECT ordered_total, real_hours, desviacion_pct
        FROM project_snapshot
        WHERE project_id = %s
        ORDER BY snapshot_year DESC, snapshot_week DESC, snapshot_at DESC
        LIMIT 1
        """,
        (project_id,),
    )
    latest_snapshot = cur.fetchone()
    ordered_total = latest_snapshot[0] if latest_snapshot else None
    real_hours = latest_snapshot[1] if latest_snapshot else None
    desviacion_pct = latest_snapshot[2] if latest_snapshot else None

    cur.execute(
        """
        INSERT INTO projects_historical (
            project_code, project_name, client, company, team, project_manager,
            consultant, status, moved_to_historical_week, progress_w,
            ordered_total, real_hours, desviacion_pct,
            moved_to_historical_at, last_import_filename
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 100, %s, %s, %s, now(), %s)
        ON CONFLICT (project_code)
        DO UPDATE SET
            project_name = EXCLUDED.project_name,
            client = EXCLUDED.client,
            company = EXCLUDED.company,
            team = EXCLUDED.team,
            project_manager = EXCLUDED.project_manager,
            consultant = EXCLUDED.consultant,
            status = EXCLUDED.status,
            moved_to_historical_week = EXCLUDED.moved_to_historical_week,
            progress_w = 100,
            ordered_total = EXCLUDED.ordered_total,
            real_hours = EXCLUDED.real_hours,
            desviacion_pct = EXCLUDED.desviacion_pct,
            moved_to_historical_at = now(),
            last_import_filename = EXCLUDED.last_import_filename
        """,
        (
            project_fields.get("project_code"),
            project_fields.get("project_name"),
            project_fields.get("client"),
            project_fields.get("company"),
            project_fields.get("team"),
            project_fields.get("project_manager"),
            project_fields.get("consultant"),
            project_fields.get("status"),
            moved_to_historical_week,
            ordered_total,
            real_hours,
            desviacion_pct,
            filename,
        ),
    )


def restore_project_from_historical(cur: psycopg.Cursor, project_code: str) -> None:
    cur.execute(
        """
        UPDATE projects
        SET is_historical = FALSE
        WHERE project_code = %s
        """,
        (project_code,),
    )
    cur.execute(
        """
        DELETE FROM projects_historical
        WHERE project_code = %s
        """,
        (project_code,),
    )

PHASES_INFO = (
    ("date_kickoff", "Kick-off"),
    ("date_design", "Design"),
    ("date_validation", "Validation"),
    ("date_golive", "Go-live"),
    ("date_reception", "Reception"),
    ("date_end", "End"),
)


def _indicator_to_color(status: str) -> str:
    if status == "red":
        return "red"
    if status in {"orange", "amber"}:
        return "orange"
    return "green"


def compute_productivity_indicator(weekly: list[dict]) -> str:
    if len(weekly) < 2:
        return "orange"
    latest = weekly[-1]
    prev = weekly[-2]

    latest_real = to_float(latest.get("real_hours"))
    prev_real = to_float(prev.get("real_hours"))
    latest_progress = to_float(latest.get("progress_w"))
    prev_progress = to_float(prev.get("progress_w"))
    latest_theoretical = to_float(latest.get("horas_teoricas"))
    prev_theoretical = to_float(prev.get("horas_teoricas"))

    if (
        latest_real is None
        or prev_real is None
        or latest_progress is None
        or prev_progress is None
    ):
        return "orange"

    if latest_real > prev_real and latest_progress <= prev_progress:
        return "red"

    if latest_theoretical is not None and latest_real > latest_theoretical:
        return "amber"

    if prev_theoretical is not None and latest_real < prev_theoretical:
        return "green"

    return "orange"


def compute_deviation_indicator(weekly: list[dict]) -> str:
    if len(weekly) < 2:
        return "orange"
    latest = weekly[-1]
    prev = weekly[-2]

    latest_dev = to_float(latest.get("desviacion_pct"))
    prev_dev = to_float(prev.get("desviacion_pct"))
    if latest_dev is None or prev_dev is None:
        return "orange"
    if latest_dev > prev_dev:
        return "red"
    if latest_dev == prev_dev:
        return "amber"
    return "green"


def compute_phase_indicator(phases_history: list[dict]) -> str:
    changes: dict[str, dict[str, bool]] = {}
    for key, _label in PHASES_INFO:
        changes[key] = {"later": False, "earlier": False, "changed": False}

    for i in range(1, len(phases_history)):
        prev = phases_history[i - 1]
        curr = phases_history[i]
        for key, _label in PHASES_INFO:
            prev_date = prev.get(key)
            curr_date = curr.get(key)
            if prev_date != curr_date:
                changes[key]["changed"] = True
                direction = "unknown"
                if prev_date and curr_date:
                    if curr_date > prev_date:
                        direction = "later"
                    elif curr_date < prev_date:
                        direction = "earlier"
                if direction == "later":
                    changes[key]["later"] = True
                if direction == "earlier":
                    changes[key]["earlier"] = True

    for key, _label in PHASES_INFO:
        if changes[key]["later"]:
            return "red"
    for key, _label in PHASES_INFO:
        if changes[key]["earlier"]:
            return "green"
    for key, _label in PHASES_INFO:
        if changes[key]["changed"]:
            return "orange"
    return "orange"


# ---------- WEB ----------
@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/estado-proyecto", response_class=HTMLResponse)
def estado_proyecto(request: Request):
    return templates.TemplateResponse("project.html", {"request": request})


@app.get("/consultas", response_class=HTMLResponse)
def consultas(request: Request):
    supported_queries = {"desviaciones": "Desviaciones"}
    consulta = request.query_params.get("consulta")
    if consulta not in supported_queries:
        consulta = None
    applied = request.query_params.get("applied") == "1"

    selected_teams = [
        value for value in request.query_params.getlist("equipo") if value
    ]
    selected_phases = [
        value for value in request.query_params.getlist("order_phase") if value
    ]

    if consulta == "desviaciones" and applied:
        results, columns, numeric_columns, row_styles = fetch_deviations_results(
            selected_teams, selected_phases
        )
    else:
        results, columns, numeric_columns, row_styles = [], [], set(), []

    teams, phases = fetch_filter_options()
    export_params = {"consulta": consulta} if consulta else {}
    if selected_teams:
        export_params["equipo"] = selected_teams
    if selected_phases:
        export_params["order_phase"] = selected_phases
    export_url = "/consultas/export"
    if export_params:
        export_url = f"{export_url}?{urllib.parse.urlencode(export_params, doseq=True)}"

    return templates.TemplateResponse(
        "queries.html",
        {
            "request": request,
            "results": results,
            "columns": columns,
            "numeric_columns": numeric_columns,
            "teams": teams,
            "phases": phases,
            "selected_teams": selected_teams,
            "selected_phases": selected_phases,
            "selected_query": consulta,
            "supported_queries": supported_queries,
            "export_url": export_url,
            "show_results": consulta == "desviaciones" and applied,
            "row_styles": row_styles,
        },
    )


@app.get("/consultas/export")
def consultas_export(request: Request):
    consulta = request.query_params.get("consulta", "desviaciones")
    selected_teams = [
        value for value in request.query_params.getlist("equipo") if value
    ]
    selected_phases = [
        value for value in request.query_params.getlist("order_phase") if value
    ]
    if consulta != "desviaciones":
        raise HTTPException(status_code=400, detail="Consulta no soportada")

    results, columns, _numeric_columns, row_styles = fetch_deviations_results(
        selected_teams, selected_phases
    )
    df = pd.DataFrame(results, columns=columns)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Consultas")
        worksheet = writer.sheets["Consultas"]
        header_fill = PatternFill("solid", fgColor="E5E7EB")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(border_style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_idx, col_name in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            max_len = len(str(col_name))
            for value in df[col_name].astype(str).tolist():
                max_len = max(max_len, len(value))
            if col_name == "Comentario":
                width = min(max_len + 2, 60)
            else:
                width = min(max_len + 2, 24)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = max(10, width)

        comment_idx = columns.index("Comentario") + 1 if "Comentario" in columns else None
        for row_idx in range(2, len(df) + 2):
            excel_row_style = row_styles[row_idx - 2] if row_idx - 2 < len(row_styles) else None
            fill = None
            if excel_row_style == "danger":
                fill = PatternFill("solid", fgColor="F8D7DA")
            elif excel_row_style == "warning":
                fill = PatternFill("solid", fgColor="FFF3CD")
            elif excel_row_style == "success":
                fill = PatternFill("solid", fgColor="D1E7DD")
            for col_idx in range(1, len(columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = border
                if fill is not None:
                    cell.fill = fill
                if col_idx == comment_idx:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
    buffer.seek(0)
    filename = "consultas_desviaciones.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers=headers,
    )


@app.get("/projects/{project_code}/indicators", response_class=HTMLResponse)
def project_indicators(request: Request, project_code: str):
    project_name = None
    try:
        with psycopg.connect(DB_DSN) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT project_name
                    FROM projects
                    WHERE project_code = %s
                      AND COALESCE(is_historical, FALSE) = FALSE
                    """,
                    (project_code,),
                )
                row = cur.fetchone()
                if row:
                    project_name = row[0]
    except Exception:
        project_name = None
    return templates.TemplateResponse(
        "indicators.html",
        {"request": request, "project_code": project_code, "project_name": project_name},
    )

@app.get("/importacion", response_class=HTMLResponse)
def importacion(request: Request):
    return templates.TemplateResponse("import.html", {"request": request})

@app.get("/menu-personal", response_class=HTMLResponse)
def menu_personal(request: Request):
    pm_name = "Alvaro Blanco Pérez"
    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT p.id,
                       p.project_code,
                       p.project_name,
                       p.client,
                       p.team,
                       s.ordered_total,
                       s.ordered_n,
                       s.ordered_e,
                       s.real_hours,
                       s.desviacion_pct,
                       s.progress_w,
                       s.payment_inv
                FROM projects p
                LEFT JOIN LATERAL (
                    SELECT ordered_total,
                           ordered_n,
                           ordered_e,
                           real_hours,
                           desviacion_pct,
                           progress_w,
                           payment_inv
                    FROM project_snapshot
                    WHERE project_id = p.id
                    ORDER BY snapshot_year DESC, snapshot_week DESC, snapshot_at DESC
                    LIMIT 1
                ) s ON TRUE
                WHERE p.project_manager = %s
                  AND COALESCE(p.is_historical, FALSE) = FALSE
                """,
                (pm_name,),
            )
            rows = cur.fetchall()

    projects = []
    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            for row in rows:
                project_id = row[0]
                cur.execute(
                    """
                    SELECT progress_w, desviacion_pct, real_hours, horas_teoricas
                    FROM project_snapshot
                    WHERE project_id = %s
                    ORDER BY snapshot_year ASC, snapshot_week ASC
                    """,
                    (project_id,),
                )
                weekly_rows = cur.fetchall()
                weekly = [
                    {
                        "progress_w": r[0],
                        "desviacion_pct": r[1],
                        "real_hours": r[2],
                        "horas_teoricas": r[3],
                    }
                    for r in weekly_rows
                ]

                cur.execute(
                    """
                    SELECT date_kickoff, date_design, date_validation,
                           date_golive, date_reception, date_end
                    FROM project_snapshot
                    WHERE project_id = %s
                    ORDER BY snapshot_year ASC, snapshot_week ASC
                    """,
                    (project_id,),
                )
                phase_rows = cur.fetchall()
                phases_history = [
                    {
                        "date_kickoff": r[0],
                        "date_design": r[1],
                        "date_validation": r[2],
                        "date_golive": r[3],
                        "date_reception": r[4],
                        "date_end": r[5],
                    }
                    for r in phase_rows
                ]

                productivity_status = compute_productivity_indicator(weekly)
                deviation_status = compute_deviation_indicator(weekly)
                phase_status = compute_phase_indicator(phases_history)
                indicator_statuses = [
                    _indicator_to_color(productivity_status),
                    _indicator_to_color(deviation_status),
                    _indicator_to_color(phase_status),
                ]
                if "red" in indicator_statuses:
                    overall_status = "red"
                elif "orange" in indicator_statuses:
                    overall_status = "orange"
                else:
                    overall_status = "green"

                ordered_total = row[5]
                if ordered_total is None and (row[6] is not None or row[7] is not None):
                    ordered_total = (row[6] or 0) + (row[7] or 0)
                projects.append(
                    {
                        "project_code": row[1],
                        "project_name": row[2],
                        "client": row[3],
                        "team": row[4],
                        "ordered_total": ordered_total,
                        "real_hours": row[8],
                        "desviacion_pct": row[9],
                        "progress_w": row[10],
                        "payment_inv": row[11],
                        "indicator_status": overall_status,
                    }
                )

    projects = sorted(projects, key=lambda item: (item["project_name"] or "").lower())

    return templates.TemplateResponse(
        "menu_personal.html",
        {"request": request, "pm_name": pm_name, "projects": projects},
    )


@app.get("/tasks", response_class=HTMLResponse)
def tasks_view(request: Request):
    return templates.TemplateResponse("tasks.html", {"request": request})


@app.get("/historicals", response_class=HTMLResponse)
def historicals(request: Request, q: str = Query("")):
    query = (q or "").strip()
    query_like = f"%{query}%"
    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            ensure_historical_storage(cur)
            cur.execute(
                """
                SELECT h.project_code,
                       h.project_name,
                       h.client,
                       h.team,
                       h.project_manager,
                       h.moved_to_historical_week,
                       h.progress_w,
                       h.ordered_total,
                       h.real_hours,
                       h.desviacion_pct
                FROM projects_historical h
                WHERE (%s = '' OR h.project_code ILIKE %s OR h.project_name ILIKE %s)
                ORDER BY h.moved_to_historical_week DESC, h.project_name ASC
                """,
                (query, query_like, query_like),
            )
            rows = cur.fetchall()

    projects = [
        {
            "project_code": r[0],
            "project_name": r[1],
            "client": r[2],
            "team": r[3],
            "project_manager": r[4],
            "moved_to_historical_week": r[5],
            "progress_w": float(r[6]) if r[6] is not None else None,
            "ordered_total": float(r[7]) if r[7] is not None else None,
            "real_hours": float(r[8]) if r[8] is not None else None,
            "desviacion_pct": float(r[9]) if r[9] is not None else None,
        }
        for r in rows
    ]
    return templates.TemplateResponse(
        "historicals.html",
        {"request": request, "projects": projects, "q": query},
    )


# ---------- API: Import ----------
@app.post("/imports")
async def import_excel(
    file: UploadFile = File(...),
    snapshot_year: int = Form(...),
    snapshot_week: int = Form(...),
    sheet: str = Form(""),
    mapping_version: str = Form(""),
    import_type: str = Form("OTS"),
):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")

    import_type = (import_type or "OTS").strip().upper()
    if import_type not in {"OTS", "ALL"}:
        raise HTTPException(status_code=400, detail="Import type must be OTS or ALL")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
        content = await file.read()
        tmp.write(content)

    df = read_and_normalize_excel(tmp_path, sheet)

    imported = 0
    skipped = 0
    archived = 0
    restored = 0

    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            ensure_historical_storage(cur)
            cur.execute(
                """
                INSERT INTO import_file (filename, snapshot_year, snapshot_week, mapping_version)
                VALUES (%s, %s, %s, %s)
                RETURNING id;
                """,
                (file.filename, snapshot_year, snapshot_week, mapping_version),
            )
            import_file_id = cur.fetchone()[0]

            for _, row in df.iterrows():
                project_fields, snapshot_fields = map_row(row)

                code = (project_fields.get("project_code") or "").strip()
                name = project_fields.get("project_name")

                if (name is None) or (str(name).strip().lower() in ("", "nan", "none")):
                    project_fields["project_name"] = f"(SIN NOMBRE) {code}"

                if not code:
                    skipped += 1
                    continue

                if import_type == "OTS":
                    pid = upsert_project(cur, project_fields)
                    snapshot_fields = compute_deltas(cur, pid, snapshot_year, snapshot_week, snapshot_fields)
                    upsert_snapshot(cur, pid, import_file_id, snapshot_year, snapshot_week, snapshot_fields)
                    imported += 1
                    continue

                internal_status = str(snapshot_fields.get("internal_status") or "").strip().lower()
                moved_to_historical_week = historical_week_label(snapshot_year, snapshot_week)
                pid = upsert_project(cur, project_fields)

                if internal_status in {"closed", "hided"}:
                    move_project_to_historical(
                        cur,
                        pid,
                        project_fields,
                        moved_to_historical_week,
                        file.filename,
                    )
                    archived += 1
                    continue

                if internal_status == "normal":
                    cur.execute(
                        """
                        SELECT 1
                        FROM projects_historical
                        WHERE project_code = %s
                        """,
                        (code,),
                    )
                    historical_row = cur.fetchone()
                    if historical_row:
                        restore_project_from_historical(cur, code)
                        restored += 1
                        snapshot_fields = compute_deltas(cur, pid, snapshot_year, snapshot_week, snapshot_fields)
                        upsert_snapshot(cur, pid, import_file_id, snapshot_year, snapshot_week, snapshot_fields)
                        imported += 1
                        continue

                    skipped += 1
                    continue

                skipped += 1

        conn.commit()

    return {
        "status": "ok",
        "import_type": import_type,
        "imported_rows": imported,
        "archived_rows": archived,
        "restored_rows": restored,
        "skipped_rows": skipped,
        "sheet": sheet,
    }


# ---------- API: Search (opcional, por si luego quieres autocompletar) ----------
@app.get("/projects/search")
def search_projects(q: str = Query(..., min_length=1), limit: int = 20):
    q_like = f"%{q}%"
    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, project_code, project_name, client, team, status
                FROM projects
                WHERE COALESCE(is_historical, FALSE) = FALSE
                  AND (project_code ILIKE %s
                   OR project_name ILIKE %s
                   OR COALESCE(client,'') ILIKE %s)
                ORDER BY project_code
                LIMIT %s
                """,
                (q_like, q_like, q_like, limit),
            )
            rows = cur.fetchall()

    return [
        {
            "id": r[0],
            "project_code": r[1],
            "project_name": r[2],
            "client": r[3],
            "team": r[4],
            "status": r[5],
        }
        for r in rows
    ]


# ---------- API: Project state ----------
@app.get("/projects/{project_code}/state")
def project_state(project_code: str, weeks_back: int = 20):
    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, project_code, project_name, client, company, team, project_manager, consultant, status
                FROM projects
                WHERE project_code = %s
                  AND COALESCE(is_historical, FALSE) = FALSE
                """,
                (project_code,),
            )
            p = cur.fetchone()
            if not p:
                raise HTTPException(status_code=404, detail="Project not found")

            project_id = p[0]

            cur.execute(
                """
                SELECT *
                FROM project_snapshot
                WHERE project_id = %s
                ORDER BY snapshot_year DESC, snapshot_week DESC, snapshot_at DESC
                LIMIT 1
                """,
                (project_id,),
            )
            latest = cur.fetchone()
            if not latest:
                raise HTTPException(status_code=404, detail="No snapshots for project")

            colnames = [desc[0] for desc in cur.description]
            latest_dict = dict(zip(colnames, latest))

            cur.execute(
                """
                SELECT snapshot_year, snapshot_week, snapshot_at,
                       progress_c, deviation_cd, payment_pending,
                       dist_c, dist_pm, dist_e
                FROM project_snapshot
                WHERE project_id = %s
                ORDER BY snapshot_year DESC, snapshot_week DESC, snapshot_at DESC
                LIMIT %s
                """,
                (project_id, weeks_back),
            )
            series_rows = cur.fetchall()

    series = [
        {
            "year": r[0],
            "week": r[1],
            "snapshot_at": r[2].isoformat() if r[2] else None,
            "progress_c": float(r[3]) if r[3] is not None else None,
            "deviation_cd": float(r[4]) if r[4] is not None else None,
            "payment_pending": float(r[5]) if r[5] is not None else None,
            "dist_c": float(r[6]) if r[6] is not None else None,
            "dist_pm": float(r[7]) if r[7] is not None else None,
            "dist_e": float(r[8]) if r[8] is not None else None,
        }
        for r in series_rows
    ]
    series = list(reversed(series))

    project = {
        "id": p[0],
        "project_code": p[1],
        "project_name": p[2],
        "client": p[3],
        "company": p[4],
        "team": p[5],
        "project_manager": p[6],
        "consultant": p[7],
        "status": p[8],
    }

    return {"project": project, "latest": latest_dict, "series": series}


# ---------- API: Project details ----------
@app.get("/projects/{project_code}/details")
def project_details(project_code: str):
    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            ensure_details_columns(cur)
            cur.execute(
                """
                SELECT id, project_code, project_name, client, company, team, project_manager, consultant, status,
                       comments,
                       hours_design, hours_development, hours_pem, hours_hypercare,
                       hours_pm, hours_consultant, hours_technician
                FROM projects
                WHERE project_code = %s
                  AND COALESCE(is_historical, FALSE) = FALSE
                """,
                (project_code,),
            )
            p = cur.fetchone()
            if not p:
                raise HTTPException(status_code=404, detail="Project not found")

            project_id = p[0]

            cur.execute(
                """
                SELECT *
                FROM project_snapshot
                WHERE project_id = %s
                ORDER BY snapshot_year DESC, snapshot_week DESC, snapshot_at DESC
                LIMIT 1
                """,
                (project_id,),
            )
            latest = cur.fetchone()
            if not latest:
                raise HTTPException(status_code=404, detail="No snapshots for project")

            colnames = [desc[0] for desc in cur.description]
            latest_dict = dict(zip(colnames, latest))

            cur.execute(
                """
                SELECT comments
                FROM project_snapshot
                WHERE project_id = %s
                  AND comments IS NOT NULL
                  AND comments <> ''
                ORDER BY snapshot_year DESC, snapshot_week DESC, snapshot_at DESC
                LIMIT 1
                """,
                (project_id,),
            )
            latest_comment_row = cur.fetchone()

    assigned_hours_phase = {
        "design": float(p[10] or 0),
        "development": float(p[11] or 0),
        "pem": float(p[12] or 0),
        "hypercare": float(p[13] or 0),
    }

    assigned_hours_role = {
        "pm": float(p[14] or 0),
        "consultant": float(p[15] or 0),
        "technician": float(p[16] or 0),
    }

    project = {
        "id": p[0],
        "project_code": p[1],
        "project_name": p[2],
        "client": p[3],
        "company": p[4],
        "team": p[5],
        "project_manager": p[6],
        "consultant": p[7],
        "status": p[8],
    }

    project_comment = p[9] if p[9] is not None else latest_dict.get("comments")
    excel_comments = latest_comment_row[0] if latest_comment_row else latest_dict.get("comments")
    return {
        "project": project,
        "latest": latest_dict,
        "assigned_hours_phase": assigned_hours_phase,
        "assigned_hours_role": assigned_hours_role,
        "project_comment": normalize_comment(project_comment),
        "excel_comments": normalize_comment(excel_comments),
    }


@app.post("/project-tasks")
def create_project_task(payload: ProjectTaskCreateIn):
    task_type = (payload.type or "").strip().upper()
    owner_role = (payload.owner_role or "").strip().upper()
    status = (payload.status or "OPEN").strip().upper()
    description = (payload.description or "").strip()

    if task_type not in TASK_TYPES:
        raise HTTPException(status_code=400, detail="Invalid task type")
    if owner_role not in TASK_OWNER_ROLES:
        raise HTTPException(status_code=400, detail="Invalid owner role")
    if status not in TASK_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid status")
    if not description:
        raise HTTPException(status_code=400, detail="Description is required")

    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            ensure_project_tasks_storage(cur)
            cur.execute(
                """
                SELECT id
                FROM projects
                WHERE id = %s
                  AND COALESCE(is_historical, FALSE) = FALSE
                """,
                (payload.project_id,),
            )
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Project not found")

            cur.execute(
                """
                INSERT INTO project_tasks (project_id, type, owner_role, planned_date, status, description)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                (payload.project_id, task_type, owner_role, payload.planned_date, status, description),
            )
            row = cur.fetchone()
        conn.commit()
    return {"id": row[0], "status": "ok"}


@app.get("/project-tasks")
def list_project_tasks(
    project_id: int | None = None,
    include_closed: bool = False,
    task_type: str | None = Query(default=None, alias="type"),
):
    where = ["COALESCE(p.is_historical, FALSE) = FALSE"]
    params: list[object] = []
    if project_id is not None:
        where.append("t.project_id = %s")
        params.append(project_id)

    normalized_type = (task_type or "").strip().upper()
    if normalized_type:
        if normalized_type not in TASK_TYPES:
            raise HTTPException(status_code=400, detail="Invalid task type")
        where.append("t.type = %s")
        params.append(normalized_type)

    if not include_closed:
        where.append("t.status <> 'CLOSED'")

    where_sql = " AND ".join(where)
    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            ensure_project_tasks_storage(cur)
            cur.execute(
                f"""
                SELECT t.id, t.project_id, p.project_code, p.project_name,
                       t.type, t.owner_role, t.planned_date, t.status, t.description,
                       t.created_at, t.updated_at
                FROM project_tasks t
                JOIN projects p ON p.id = t.project_id
                WHERE {where_sql}
                ORDER BY t.created_at DESC, t.id DESC
                """,
                params,
            )
            rows = cur.fetchall()
    return [
        {
            "id": r[0],
            "project_id": r[1],
            "project_code": r[2],
            "project_name": r[3],
            "type": r[4],
            "owner_role": r[5],
            "planned_date": to_date_iso(r[6]),
            "status": r[7],
            "description": r[8],
            "created_at": r[9].isoformat() if r[9] else None,
            "updated_at": r[10].isoformat() if r[10] else None,
        }
        for r in rows
    ]


@app.patch("/project-tasks/{task_id}/status")
def update_project_task_status(task_id: int, payload: ProjectTaskStatusIn):
    status = (payload.status or "").strip().upper()
    if status not in TASK_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid status")

    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            ensure_project_tasks_storage(cur)
            cur.execute(
                """
                UPDATE project_tasks t
                SET status = %s,
                    updated_at = now()
                FROM projects p
                WHERE t.id = %s
                  AND p.id = t.project_id
                  AND COALESCE(p.is_historical, FALSE) = FALSE
                RETURNING t.id
                """,
                (status, task_id),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Task not found")
        conn.commit()
    return {"status": "ok"}


@app.put("/project-tasks/{task_id}")
def update_project_task(task_id: int, payload: ProjectTaskUpdateIn):
    task_type = (payload.type or "").strip().upper()
    owner_role = (payload.owner_role or "").strip().upper()
    status = (payload.status or "").strip().upper()
    description = (payload.description or "").strip()

    if task_type not in TASK_TYPES:
        raise HTTPException(status_code=400, detail="Invalid task type")
    if owner_role not in TASK_OWNER_ROLES:
        raise HTTPException(status_code=400, detail="Invalid owner role")
    if status not in TASK_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid status")
    if not description:
        raise HTTPException(status_code=400, detail="Description is required")

    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            ensure_project_tasks_storage(cur)
            cur.execute(
                """
                UPDATE project_tasks t
                SET type = %s,
                    owner_role = %s,
                    planned_date = %s,
                    status = %s,
                    description = %s,
                    updated_at = now()
                FROM projects p
                WHERE t.id = %s
                  AND p.id = t.project_id
                  AND COALESCE(p.is_historical, FALSE) = FALSE
                RETURNING t.id
                """,
                (task_type, owner_role, payload.planned_date, status, description, task_id),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Task not found")
        conn.commit()
    return {"status": "ok"}


@app.get("/projects/{project_code}/task-counters")
def project_task_counters(project_code: str):
    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            ensure_project_tasks_storage(cur)
            cur.execute(
                """
                SELECT p.id
                FROM projects p
                WHERE p.project_code = %s
                  AND COALESCE(p.is_historical, FALSE) = FALSE
                """,
                (project_code,),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Project not found")
            project_id = row[0]

            cur.execute(
                """
                SELECT type, COUNT(*)
                FROM project_tasks
                WHERE project_id = %s
                  AND status <> 'CLOSED'
                GROUP BY type
                """,
                (project_id,),
            )
            counts = {r[0]: int(r[1]) for r in cur.fetchall()}

    return {
        "project_id": project_id,
        "task_open_count": counts.get("TASK", 0),
        "pp_open_count": counts.get("PP", 0),
    }


@app.get("/projects/{project_code}/metrics/weekly")
def project_weekly_metrics(project_code: str):
    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id
                FROM projects
                WHERE project_code = %s
                  AND COALESCE(is_historical, FALSE) = FALSE
                """,
                (project_code,),
            )
            project_row = cur.fetchone()
            if not project_row:
                raise HTTPException(status_code=404, detail="Project not found")
            project_id = project_row[0]

            cur.execute(
                """
                SELECT snapshot_year, snapshot_week,
                       progress_w, desviacion_pct,
                       real_hours, horas_teoricas,
                       progress_w_delta, real_hours_delta,
                       horas_teoricas_delta, desviacion_pct_delta
                FROM project_snapshot
                WHERE project_id = %s
                ORDER BY snapshot_year ASC, snapshot_week ASC
                """,
                (project_id,),
            )
            rows = cur.fetchall()

    return [
        {
            "year": r[0],
            "week": r[1],
            "progress_w": to_float(r[2]),
            "desviacion_pct": to_float(r[3]),
            "real_hours": to_float(r[4]),
            "horas_teoricas": to_float(r[5]),
            "progress_w_delta": to_float(r[6]),
            "real_hours_delta": to_float(r[7]),
            "horas_teoricas_delta": to_float(r[8]),
            "desviacion_pct_delta": to_float(r[9]),
        }
        for r in rows
    ]


@app.get("/projects/{project_code}/metrics/phases")
def project_phase_history(project_code: str):
    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id
                FROM projects
                WHERE project_code = %s
                  AND COALESCE(is_historical, FALSE) = FALSE
                """,
                (project_code,),
            )
            project_row = cur.fetchone()
            if not project_row:
                raise HTTPException(status_code=404, detail="Project not found")
            project_id = project_row[0]

            cur.execute(
                """
                SELECT snapshot_year, snapshot_week,
                       date_kickoff, date_design, date_validation,
                       date_golive, date_reception, date_end
                FROM project_snapshot
                WHERE project_id = %s
                ORDER BY snapshot_year ASC, snapshot_week ASC
                """,
                (project_id,),
            )
            rows = cur.fetchall()

    return [
        {
            "year": r[0],
            "week": r[1],
            "date_kickoff": to_date_iso(r[2]),
            "date_design": to_date_iso(r[3]),
            "date_validation": to_date_iso(r[4]),
            "date_golive": to_date_iso(r[5]),
            "date_reception": to_date_iso(r[6]),
            "date_end": to_date_iso(r[7]),
        }
        for r in rows
    ]


@app.post("/projects/{project_id}/assigned-hours/phase")
def update_assigned_hours_phase(project_id: int, payload: AssignedHoursPhaseIn):
    if payload.phase not in PHASES:
        raise HTTPException(status_code=400, detail="Invalid phase")
    hours = payload.hours if payload.hours is not None else 0
    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            ensure_details_columns(cur)
            column_map = {
                "design": "hours_design",
                "development": "hours_development",
                "pem": "hours_pem",
                "hypercare": "hours_hypercare",
            }
            column = column_map[payload.phase]
            cur.execute(
                f"""
                UPDATE projects
                SET {column} = %s
                WHERE id = %s
                """,
                (hours, project_id),
            )
        conn.commit()
    return {"status": "ok"}


@app.post("/projects/{project_id}/assigned-hours/role")
def update_assigned_hours_role(project_id: int, payload: AssignedHoursRoleIn):
    if payload.role not in ROLES:
        raise HTTPException(status_code=400, detail="Invalid role")
    hours = payload.hours if payload.hours is not None else 0
    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            ensure_details_columns(cur)
            column_map = {
                "pm": "hours_pm",
                "consultant": "hours_consultant",
                "technician": "hours_technician",
            }
            column = column_map[payload.role]
            cur.execute(
                f"""
                UPDATE projects
                SET {column} = %s
                WHERE id = %s
                """,
                (hours, project_id),
            )
        conn.commit()
    return {"status": "ok"}


@app.post("/projects/{project_id}/comments")
def update_project_comment(project_id: int, payload: ProjectCommentIn):
    with psycopg.connect(DB_DSN) as conn:
        with conn.cursor() as cur:
            ensure_details_columns(cur)
            cur.execute(
                """
                UPDATE projects
                SET comments = %s
                WHERE id = %s
                """,
                (payload.comment_text, project_id),
            )
        conn.commit()
    return {"status": "ok"}
